import os
import shutil
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def trouver_colonne_date(df):
    for col in df.columns:
        if isinstance(col, str) and 'DATE' in col.upper():
            # Essayer de convertir la colonne en datetime
            try:
                df[col] = pd.to_datetime(df[col], format='%Y-%m-%d %H:%M:%S')
                return col
            except:
                pass
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            return col
    return None

def charger_fichier_excel(chemin_fichier):
    # Lire tout le fichier Excel sans sauter de lignes
    df = pd.read_excel(chemin_fichier, header=0)  # Utiliser la première ligne comme en-tête
    
    print(f"Nombre de colonnes dans le fichier Excel : {df.shape[1]}")
    print("Premières lignes du DataFrame original:")
    print(df.head())
    print("Types des colonnes:")
    print(df.dtypes)
    
    date_column = trouver_colonne_date(df)
    if date_column:
        print(f"Colonne de date trouvée : {date_column}")
        return df.dropna(subset=[date_column], how='all')
    else:
        print("Aucune colonne de date trouvée.")
        return df

def creer_dossier_si_non_existant(dossier):
    os.makedirs(dossier, exist_ok=True)

def deplacer_fichier(source, destination):
    shutil.move(source, destination)

def traiter_fichiers(dossier_a_faire, dossier_fait, dossier_archive):
    for dossier in [dossier_a_faire, dossier_fait, dossier_archive]:
        creer_dossier_si_non_existant(dossier)

    for fichier in os.listdir(dossier_a_faire):
        if fichier.endswith('.xlsx'):
            chemin_fichier = os.path.join(dossier_a_faire, fichier)
            df = charger_fichier_excel(chemin_fichier)
            
            date_column = trouver_colonne_date(df)
            if date_column is None:
                print(f"Aucune colonne de date trouvée dans le fichier {fichier}. Passons au fichier suivant.")
                continue
            
            session = os.path.splitext(fichier)[0]
            date_debut = df[date_column].min().date()
            date_fin = df[date_column].max().date()
            # Generate the new filename
            nouveau_nom_fichier = f"planning YP - {session} - du {datetime.now().strftime('%d-%m-%Y')}.xlsx"
            chemin_fichier_sortie = os.path.join(dossier_fait, nouveau_nom_fichier)
            
            traiter_session(df, session, date_debut, date_fin, fichier)
            # Rename and move the file
            os.rename(os.path.join(dossier_fait, f'planning YP  - du {datetime.now().strftime("%d-%m-%Y")}.xlsx'), chemin_fichier_sortie)
            # deplacer_fichier(chemin_fichier, os.path.join(dossier_archive, fichier))
            
def traitement(df, date_column):
    colonnes = list(df.columns)
    matiere_col = colonnes[1] if len(colonnes) > 1 else None
    formateur_col = colonnes[2] if len(colonnes) > 2 else None
    type_cours_col = colonnes[3] if len(colonnes) > 3 else None

    # Trier le DataFrame par date
    df = df.sort_values(by=date_column)

    donnees_traitees = []

    for _, row in df.iterrows():
        donnees_traitees.append({
            'StartDate': row[date_column].date(),
            'EndDate': row[date_column].date(),
            'Matiere': row[matiere_col] if matiere_col else 'Inconnue',
            'Formateur': row[formateur_col] if formateur_col else 'Inconnu',
            'Modalite': row[type_cours_col] if type_cours_col else 'Inconnu',
            'Count': 1
        })

    return donnees_traitees
def fusionner_lignes_consecutives(donnees_traitees):
    if not donnees_traitees:
        return []

    resultat = [donnees_traitees[0]]

    for ligne_actuelle in donnees_traitees[1:]:
        ligne_precedente = resultat[-1]
        
      
        days_difference = (ligne_actuelle['StartDate'] - ligne_precedente['EndDate']).days
        
       
        if (ligne_actuelle['Matiere'] == ligne_precedente['Matiere'] and
            ligne_actuelle['Formateur'] == ligne_precedente['Formateur'] and
            ligne_actuelle['Modalite'] == ligne_precedente['Modalite'] and
            days_difference <= 4):  
            
            ligne_precedente['EndDate'] = ligne_actuelle['EndDate']
            ligne_precedente['Count'] += ligne_actuelle['Count']
        else:
            resultat.append(ligne_actuelle)

    return resultat


          
def traiter_session(df, session, date_debut, date_fin, fichier_original):
    date_column = trouver_colonne_date(df)
    donnees_traitees = traitement(df, date_column)
    donnees_fusionnees = fusionner_lignes_consecutives(donnees_traitees)

    data = []
    for ligne in donnees_fusionnees:
        start_date = ligne['StartDate'].strftime('%d-%m-%Y')
        end_date = ligne['EndDate'].strftime('%d-%m-%Y')
        data.append([start_date, end_date, ligne['Matiere'], ligne['Formateur'], ligne['Modalite'], ligne['Count']])

    df_result = pd.DataFrame(data, columns=['Du', 'Au', 'Matière', 'Formateur', 'Modalité', 'Nb jours'])
    df_result['Modalité'] = df_result['Modalité'].str.lower()
    df_result['Jour(s) facturables'] = df_result['Nb jours']
    df_result.loc[df_result['Modalité'] == 'autoformation', 'Jour(s) facturables'] = 0

    total_jours = df_result['Nb jours'].sum()
    total_jours_facturables = df_result['Jour(s) facturables'].sum()
    df_result.loc[len(df_result)] = ['', '', 'Total jours centre', '', '', total_jours, total_jours_facturables]
    df_result.loc[len(df_result)] = ['', '', 'Total heures centre', '', '', total_jours * 7, total_jours_facturables * 7]

    nom_fichier_sortie = f'planning YP  - du {datetime.now().strftime("%d-%m-%Y")}.xlsx'
    chemin_fichier_sortie = os.path.join('fait', nom_fichier_sortie)
    formater_excel(df_result, chemin_fichier_sortie, session, date_debut, date_fin, fichier_original)
    print(f"Planning exporté vers {chemin_fichier_sortie}")
  
 
def get_formatted_filename(fichier_original):
    session = os.path.splitext(fichier_original)[0]
    current_date_str = datetime.now().strftime('%d-%m-%Y')  
    return f"planning YP - {session} - du {current_date_str}.xlsx" 
     
def formater_excel(df, chemin_fichier_sortie, session, date_debut, date_fin, fichier_original):
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"
    ws.row_dimensions[1].height = 50

    font_calibri_10 = Font(name='Calibri', size=10)
    border_style = Border(left=Side(style='thin', color='BFBFBF'),
                          right=Side(style='thin', color='BFBFBF'),
                          top=Side(style='thin', color='BFBFBF'),
                          bottom=Side(style='thin', color='BFBFBF'))

    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    date_debut_str = date_debut.strftime('%d/%m/%Y')
    date_fin_str = date_fin.strftime('%d/%m/%Y')

    # Set the title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f'Planning {session} du {date_debut_str} au {date_fin_str}'
    title_cell.font = Font(name='Calibri', size=10, color='FFFFFF')
    title_cell.fill = PatternFill(start_color="003350", end_color="003350", fill_type="solid")
    title_cell.alignment = alignment_center

    # Headers
    headers = ['Du', 'Au', 'Matière', 'Formateur', 'Modalité', 'Nb jours', 'Jour(s) facturables']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = font_calibri_10
        cell.border = border_style
        cell.alignment = alignment_center
        cell.fill = PatternFill(start_color="003350", end_color="003350", fill_type="solid")
        cell.font = Font(name='Calibri', size=10, color='FFFFFF')

    # Add data
    for r_idx, row in enumerate(df.itertuples(), start=3):
        row_fill = None
        matiere = str(row.Matière).strip().upper()  # Convert to string, strip whitespace, and uppercase
        # Determine row fill color
        if 'ACCUEIL' in matiere or matiere == 'EXAMEN':
            row_fill = PatternFill(start_color="F9BA00", end_color="F9BA00", fill_type="solid")
        elif 'PROJET' in matiere:
            row_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        elif str(row.Modalité).lower() == 'autoformation':
            row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        for c_idx, value in enumerate(row[1:8], start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value if value != '(vide)' else '')
            cell.font = font_calibri_10
            cell.border = border_style
            cell.alignment = alignment_left if c_idx == 3 else alignment_center
            
            if row_fill:
                cell.fill = row_fill

    # Add totals rows
    total_jours = df['Nb jours'].sum()
    total_jours_facturables = df['Jour(s) facturables'].sum()

    for row in [ws.max_row - 1, ws.max_row]:
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="003350", end_color="003350", fill_type="solid")
            cell.font = Font(name='Calibri', size=10, color='FFFFFF')

   
    ws.append(['', '', '', '', '', '', ''])  
    edition_date = datetime.now().strftime('%d/%m/%Y')
    date_cell = ws.cell(row=ws.max_row, column=3, value=f"Planning édité le {edition_date}")
    date_cell.font = font_calibri_10
    date_cell.alignment = alignment_left

    # Add another blank row and "Fermeture centre"
    ws.append(['', '', '', '', '', '', ''])  
    fermeture_cell = ws.cell(row=ws.max_row, column=1, value="Fermeture centre")
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=7)
    fermeture_cell.fill = PatternFill(start_color="003350", end_color="003350", fill_type="solid")
    fermeture_cell.font = Font(name='Calibri', size=10, color='FFFFFF')
    fermeture_cell.alignment = alignment_center

    # Adjust column widths
    column_widths = [2.2, 2.2, 7.6, 2.7, 2.9, 1.3, 1.3]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width * 7

    # Add logo if exists
    logo_path = 'logo_diginamic.png'
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 150 
        img.height = 65  
        ws.add_image(img, 'A1')

    # Set margins
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25

    # Set the orientation to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Save the Excel file
    wb.save(chemin_fichier_sortie)

if __name__ == "__main__":
    dossier_a_faire = 'a-faire'
    dossier_fait = 'fait'
    dossier_archive = 'archive'
    traiter_fichiers(dossier_a_faire, dossier_fait, dossier_archive)